from __future__ import annotations

from dataclasses import dataclass
from typing import List, Dict
import re

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils.cell import range_boundaries


DAY_ALIASES: Dict[str, str] = {
	"mon": "Mon",
	"monday": "Mon",
	"dushanba": "Mon",
	"du": "Mon",
	"pon": "Mon",
	"пон": "Mon",
	"понедельник": "Mon",
	"dúyshembi": "Mon",

	"tue": "Tue",
	"tuesday": "Tue",
	"seshanba": "Tue",
	"se": "Tue",
	"vt": "Tue",
	"вт": "Tue",
	"вторник": "Tue",
	"shiyshembi": "Tue",

	"wed": "Wed",
	"wednesday": "Wed",
	"chorshanba": "Wed",
	"cho": "Wed",
	"sr": "Wed",
	"ср": "Wed",
	"среда": "Wed",
	"sárshembi": "Wed",

	"thu": "Thu",
	"thursday": "Thu",
	"payshanba": "Thu",
	"pay": "Thu",
	"чт": "Thu",
	"четверг": "Thu",
	"piyshembi": "Thu",

	"fri": "Fri",
	"friday": "Fri",
	"juma": "Fri",
	"ju": "Fri",
	"пт": "Fri",
	"пятница": "Fri",
	"juma": "Fri",

	"sat": "Sat",
	"saturday": "Sat",
	"shanba": "Sat",
	"sha": "Sat",
	"сб": "Sat",
	"суббота": "Sat",
	"shembi": "Sat"
}


@dataclass(frozen=True)
class NormalizedRow:
	group: str
	day: str
	time: str | None
	subject: str
	teacher: str
	room: str | None


def _normalize_day(value: str) -> str:
	key = (value or "").strip().lower()
	return DAY_ALIASES.get(key, value.strip())


def _find_day_in_text(value: str) -> str | None:
	text = (value or "").lower()
	for k, std in DAY_ALIASES.items():
		if k and k in text:
			return std
	return None


def _expand_merged_cells(ws: Worksheet) -> Dict[tuple[int, int], str]:
	# Map of (row, col) -> value for cells covered by merges
	# In read_only mode, merged ranges are not available; handle gracefully
	expanded: Dict[tuple[int, int], str] = {}
	merged_ranges = getattr(ws, "merged_cells", None)
	if not merged_ranges:
		return expanded
	for mr in merged_ranges.ranges:
		min_col, min_row, max_col, max_row = range_boundaries(str(mr))
		base_val = ws.cell(row=min_row, column=min_col).value
		for r in range(min_row, max_row + 1):
			for c in range(min_col, max_col + 1):
				expanded[(r, c)] = base_val
	return expanded


def _try_parse_matrix(ws: Worksheet) -> List[NormalizedRow]:
    max_row = ws.max_row or 0
    max_col = ws.max_column or 0
    merged_map = _expand_merged_cells(ws)

    def get_val(r: int, c: int) -> str:
        val = merged_map.get((r, c))
        if val is None:
            val = ws.cell(row=r, column=c).value
        return "" if val is None else str(val).strip()

    # Find group headers in row 7 (group numbers) and row 8 (group names)
    group_cols: Dict[str, int] = {}
    group_names: Dict[int, str] = {}
    
    # Get group numbers from row 7
    for c in range(1, max_col + 1):
        val = get_val(7, c)
        if val and "(" in val and ")" in val and "-" in val and val != "s/s":
            group_cols[val] = c
    
    # Get group names from row 8
    for c in range(1, max_col + 1):
        val = get_val(8, c)
        if val and val not in ["s/s"] and not val.isdigit():
            group_names[c] = val

    if not group_cols:
        return []

    rows: List[NormalizedRow] = []
    
    # Process each day block
    day_blocks = [
        (9, 16, "Mon"),   # Dúyshembi
        (17, 24, "Tue"),  # Shiyshembi  
        (25, 32, "Wed"),  # Sárshembi
        (33, 40, "Thu"),  # Piyshembi
        (41, 48, "Fri"),  # Juma
        (49, 54, "Sat"),  # Shembi
    ]
    
    # Helper to convert 1..10 to Roman numerals
    roman_map = {1: "I", 2: "II", 3: "III", 4: "IV", 5: "V", 6: "VI", 7: "VII", 8: "VIII", 9: "IX", 10: "X"}

    for start_row, end_row, day_name in day_blocks:
        for r in range(start_row, min(end_row + 1, max_row + 1)):
            time_val = get_val(r, 2).strip() or None
            # If time is missing, derive para number from row position within the day block (every two rows per para)
            if not time_val:
                offset = r - start_row
                para_num = (offset // 2) + 1
                time_val = roman_map.get(para_num, str(para_num))
            
            # Process each group individually
            for group_number, group_col in group_cols.items():
                subject_cell = get_val(r, group_col)
                teacher_cell = get_val(r + 1, group_col)
                
                if not subject_cell or not teacher_cell:
                    continue
                
                # Parse subject
                subject_lines = [ln.strip() for ln in subject_cell.split("\n") if ln.strip()]
                teacher_lines = [ln.strip() for ln in teacher_cell.split("\n") if ln.strip()]
                
                if not subject_lines:
                    continue
                    
                subject_raw = subject_lines[0]
                
                # Skip if subject is actually a teacher name
                teacher_titles = ['ass.', 'prof.', 'phd.', 'dr.', 'doc.', 'assistant', 'professor']
                subject_lower = subject_raw.lower()
                if any(subject_lower.startswith(title) for title in teacher_titles):
                    continue
                
                # Parse teacher
                teacher_raw = teacher_lines[0] if teacher_lines else ""
                teacher_clean = re.sub(r'\s*\([^)]*\)\s*$', '', teacher_raw).strip()
                
                teacher = ""
                if teacher_clean:
                    teacher_lower = teacher_clean.lower()
                    if any(teacher_lower.startswith(title) for title in teacher_titles):
                        teacher = teacher_clean
                
                # Check if this is a common subject and get room number from s/s column
                is_common_subject = False
                room = None
                
                # Find s/s column after this group
                s_s_col = group_col + 1
                if s_s_col <= max_col:
                    s_s_val = get_val(r, s_s_col)
                    
                    # If s/s column is empty, this might be a common subject
                    if not s_s_val or str(s_s_val).strip() == "":
                        # Check if next group column is also empty (common subject)
                        next_group_col = None
                        for g_num, g_col in group_cols.items():
                            if g_col > group_col:
                                next_group_col = g_col
                                break
                        
                        if next_group_col:
                            next_subject_cell = get_val(r, next_group_col)
                            if not next_subject_cell or str(next_subject_cell).strip() == "":
                                # Both current group and next group columns are empty
                                # This means it's a common subject
                                is_common_subject = True
                                
                                # For common subjects, look for room number in the next s/s column
                                next_s_s_col = next_group_col + 1
                                if next_s_s_col <= max_col:
                                    next_s_s_val = get_val(r, next_s_s_col)
                                    if next_s_s_val and str(next_s_s_val).strip().isdigit():
                                        room = str(next_s_s_val).strip()
                    elif s_s_val and str(s_s_val).strip().isdigit():
                        # This is a room number - individual subject
                        room = str(s_s_val).strip()
                        is_common_subject = False
                
                if is_common_subject:
                    # This is a common subject - add to current group and next group
                    # Find next group
                    next_group_col = None
                    for g_num, g_col in group_cols.items():
                        if g_col > group_col:
                            next_group_col = g_col
                            break
                    
                    # Add to current group
                    group_name = group_names.get(group_col, "")
                    if group_name and group_name != group_number:
                        group_display_name = f"{group_number} ({group_name})"
                    else:
                        group_display_name = group_number
                    
                    rows.append(NormalizedRow(
                        group=group_display_name, 
                        day=day_name, 
                        time=time_val, 
                        subject=subject_raw, 
                        teacher=teacher, 
                        room=room
                    ))
                    
                    # Add to next group if exists
                    if next_group_col:
                        next_group_number = None
                        for g_num, g_col in group_cols.items():
                            if g_col == next_group_col:
                                next_group_number = g_num
                                break
                        
                        if next_group_number:
                            next_group_name = group_names.get(next_group_col, "")
                            if next_group_name and next_group_name != next_group_number:
                                next_group_display_name = f"{next_group_number} ({next_group_name})"
                            else:
                                next_group_display_name = next_group_number
                            
                            rows.append(NormalizedRow(
                                group=next_group_display_name, 
                                day=day_name, 
                                time=time_val, 
                                subject=subject_raw, 
                                teacher=teacher, 
                                room=room
                            ))
                else:
                    # Individual subject
                    group_name = group_names.get(group_col, "")
                    if group_name and group_name != group_number:
                        group_display_name = f"{group_number} ({group_name})"
                    else:
                        group_display_name = group_number
                    
                    rows.append(NormalizedRow(
                        group=group_display_name, 
                        day=day_name, 
                        time=time_val, 
                        subject=subject_raw, 
                        teacher=teacher, 
                        room=room
                    ))

    return rows


# Monkey-patch: if the flat-table parser yields nothing, attempt matrix parser
def load_schedule_from_excel(file_path: str) -> List[NormalizedRow]:
	wb = load_workbook(filename=file_path, read_only=True, data_only=True)
	ws = wb.active
	# First, try flat-table format
	rows_iter = ws.iter_rows(values_only=True)
	try:
		headers = next(rows_iter)
	except StopIteration:
		headers = []

	index: Dict[str, int] = {}
	for i, h in enumerate(headers or []):
		if not h:
			continue
		n = str(h).strip().lower()
		index[n] = i

	def col(name_alts: List[str]) -> int | None:
		for n in name_alts:
			if n in index:
				return index[n]
		return None

	group_col = col(["group", "guruh", "группа", "gruppa"]) 
	day_col = col(["day", "kun", "день", "den"]) 
	time_col = col(["time", "soat", "время", "vaqt"]) 
	subject_col = col(["subject", "fan", "предмет"]) 
	teacher_col = col(["teacher", "o'qituvchi", "oqituvchi", "преподаватель", "teacher name"]) 
	room_col = col(["room", "auditoriya", "аудитория"]) 

	result: List[NormalizedRow] = []
	if headers and (group_col is not None or (day_col is not None and subject_col is not None)):
		for row in rows_iter:
			def v(idx: int | None) -> str:
				if idx is None or idx >= len(row):
					return ""
				val = row[idx]
				return "" if val is None else str(val).strip()

			group = v(group_col)
			day = _normalize_day(v(day_col))
			time = v(time_col) or None
			subject = v(subject_col)
			teacher = v(teacher_col)
			room = v(room_col) or None

			if not group or not day or not subject:
				continue
			result.append(NormalizedRow(group=group, day=day, time=time, subject=subject, teacher=teacher, room=room))

	if result:
		return result

	# Fallback to matrix parser
	rows = _try_parse_matrix(ws)
	if rows:
		return rows

	# As a last resort, dump a preview for debugging so user can share it
	try:
		from pathlib import Path
		preview_lines: List[str] = []
		max_r = min(ws.max_row or 0, 60)
		max_c = min(ws.max_column or 0, 30)
		for r in range(1, max_r + 1):
			vals = []
			for c in range(1, max_c + 1):
				cell = ws.cell(row=r, column=c).value
				vals.append("" if cell is None else str(cell).replace("\n", " "))
			preview_lines.append("\t".join(vals))
		out_dir = Path.cwd() / "data"
		out_dir.mkdir(parents=True, exist_ok=True)
		out_path = out_dir / "sheet_preview.txt"
		out_path.write_text("\n".join(preview_lines), encoding="utf-8")
	except Exception:
		pass

	return []